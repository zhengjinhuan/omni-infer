import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker

import ast
import os
import re
import sys
from datetime import datetime

action_dict = {
    'PNode': 'P节点',
    'DNode': 'D节点',
    'Start to schedule': '调度进程收到请求',
    'Finish to choose device and add request': '选择设备，准备传递请求',
    'Start to send request to pd api server': '准备发送请求给prefill api server',

    'p_1 prefill api server收到请求': 'p_1 prefill api server收到请求',
    'p_2 触发engine处理请求': 'p_2 触发engine处理请求',
    'p_3 engine开始tokennizer': 'p_3 engine开始tokennizer',
    'p_3 engine结束tokennizer': 'p_3 engine结束tokennizer',
    'p_4 tokennizer to sche': 'p_4 tokennizer to sche',
    'p_5 P侧添加到bootstrap队列之后': 'p_5 P侧添加到bootstrap队列之后',
    'P侧握手完成': 'P侧握手完成',
    'p_6 握手完成加到waiting队列': 'p_6 握手完成加到waiting队列',
    'p_7 组batch完成 开始run batch': 'p_7 组batch完成 开始run batch',
    'p_8 Push a new batch to the input queue': 'p_8 Push a new batch to the input queue',
    'p_9 开始执行Run forward': 'p_9 开始执行Run forward',
    'p_9 结束执行Run forward': 'p_9 结束执行Run forward',
    'p_13 开始发送kv cache': 'p_13 开始发送kv cache',
    'p_14 完成发送kv cache': 'p_14 完成发送kv cache',
    'p_15 P侧释放KV': 'p_15 P侧释放KV',
    'p_16 client收到输出并入队': 'p_16 client收到输出并入队',
    'p_17 client出队': 'p_17 client出队',
    'p_18 api server收到请求准备返回': 'p_18 api server收到请求准备返回',

    'Get response from pd api server': '调度进程收到api server返回的响应',
    'Finish to chosse device and start decode generate': '调度进程选择decoder device 准备发送请求',
    'Finish to decode': '调度进程完成发送',

    'waiting_pull_len': '添加后need pulling队列长度',

    'd_1 decode api server收到请求': 'd_1 decode api server收到请求',
    'd_2 触发engine处理请求': 'd_2 触发engine处理请求',
    'd_3 engine开始tokennizer': 'd_3 engine开始tokennizer',
    'd_3 engine结束tokennizer': 'd_3 engine结束tokennizer',
    'd_4 tokennizer to sche': 'd_4 tokennizer to sche',
    'd_5 scheduler开始处理请求': 'd_5 scheduler开始处理请求',
    'd_6 D侧添加到prealloc_queue队列之后': 'd_6 D侧添加到prealloc_queue队列之后',
    'd_7 Add need pullling sequence': 'd_7 Add need pullling sequence',
    'd_8 开始分配kv缓存': 'd_8 开始分配kv缓存',
    'd_9 d侧开始握手': 'd_9 d侧开始握手',
    'd_11 d侧收到kv传输完成通知': 'd_11 d侧收到kv传输完成通知',
    'd_12 轮询到状态 kv cache传输完成 开始加入到waiting队列': 'd_12 轮询到状态 kv cache传输完成 开始加入到waiting队列',
    'd_18 触发首个decode token执行': 'd_18 触发首个decode token执行',
    'd_19 decoder返回第一个token': 'd_19 decoder返回第一个token',
    'd_20 decoder返回第二个token': 'd_20 decoder返回第二个token',
    'd_21 api server收到推理结果': 'd_21 api server收到推理结果',
}


def _sgl_title() -> list:
    """ 列名 """

    titles = [
        '请求ID',  # req_id
        # 'Prefill ID',  # Prefill id
        "请求长度",  # Seq len
        "decode输出的tokens",  # decode toknes number
        "处理请求P节点ID",  # PNode
        "处理请求D节点ID",  # DNode

        '调度进程收到请求',  # Start to schedule
        '选择设备，准备传递请求',  # Finish to choose device and add request
        '准备发送请求给prefill api server',  # Start to send request to pd api server

        'p_1 prefill api server收到请求',  # PD api server get request
        'p_2 触发engine处理请求',  # Get prefill engine request and start pickle
        'p_3 engine开始tokennizer',  # Finish process request in prefill engine
        'p_3 engine结束tokennizer',  # Start process request in prefill engine
        'p_4 tokennizer to sche',  # Prefill add waiting queue
        'p_5 P侧添加到bootstrap队列之后',  # Start engine step
        'P侧握手完成',
        'p_6 握手完成加到waiting队列',  # try to schedule in waiting queue
        'p_7 组batch完成 开始run batch',  # fail to add result of kv insufficient
        'p_8 Push a new batch to the input queue',  # fail to add result of can not schedule
        'p_9 开始执行Run forward',  # Prefill get new_blocks
        'p_9 结束执行Run forward',  # Finish engine step
        'p_13 开始发送kv cache',  # Client get prefill output #0613
        'p_14 完成发送kv cache',  # Pop output queues #0613
        'p_15 P侧释放KV',  # Finish prefill pickle and start response
        'p_16 client收到输出并入队',  # Get response from pd api server
        'p_17 client出队',  # Finish to chosse device and start decode generate
        'p_18 api server收到请求准备返回',  # Finish to decode

        '调度进程收到api server返回的响应',  # Get response from pd api server
        '调度进程选择decoder device 准备发送请求',  # Finish to chosse device and start decode generate
        '调度进程完成发送',  # Finish to decode

        'd_1 decode api server收到请求',  # waiting_pull_len
        'd_2 触发engine处理请求',  # Start pull kv #0613
        'd_3 engine开始tokennizer',  # Finish pull kv #0613
        'd_3 engine结束tokennizer',  # Prefill free kv blocks
        'd_4 tokennizer to sche',  # Start append running sequece for decode
        'd_5 scheduler开始处理请求',  # Start to send output
        'd_6 D侧添加到prealloc_queue队列之后',  # First decode output token
        'd_8 开始分配kv缓存',  # Scend decode output token
        'd_9 d侧开始握手',  # Scend decode output token
        'd_7 Add need pullling sequence',
        'waiting_pull_len',  # Get first token
        'd_11 d侧收到kv传输完成通知',  # Scend decode output token
        'd_12 轮询到状态 kv cache传输完成 开始加入到waiting队列',  # Scend decode output token
        'd_18 触发首个decode token执行',  # Scend decode output token
        'd_19 decoder返回第一个token',  # Scend decode output token
        'd_20 decoder返回第二个token',  # Scend decode output token
        'd_21 api server收到推理结果'  # Finish decode pickle and start response
    ]

    return titles

def decode_worker_step(node, engine_core_str, lines_decode_worker_step):
    node_key = node + '|' + engine_core_str
    worker_step: int = None
    if node_key in lines_decode_worker_step.keys():
        worker_step = lines_decode_worker_step[node_key]
    return worker_step


def parse_file(folder_path):
    lines = []
    lines_decode_step = []
    lines_decode_worker_step = {}
    step_data = []  # req_ids : start engine time, end engine time, 同一个batch总的tokens
    step_data_decode = {}
    metric = []
    for item in os.listdir(folder_path):
        item_path = os.path.join(folder_path, item)
        if os.path.isfile(item_path):  # 过滤文件
            print(item_path)  # 解析文件
            with open(item_path, 'r', encoding='utf-8') as file:
                for line in file:
                    line = line.strip()
                    if not line:
                        continue
                    if 'CompletionMetric' in line:
                        metric.append(line)
                        continue
                    if 'profile' not in line:
                        continue
                    node_info = None
                    if '_NODE_' in item:
                        idx0 = item.find('_NODE_') + len('_NODE_')
                        idx1 = item.find('_', idx0)
                        node_info = item[idx0:idx1]
                    if 'Times' in line:
                        lines.append(line)
                    if 'engine_step start' in line:  # 含有engine_step start的文件必然有_NODE_信息
                        if node_info.startswith('P'):
                            lines.append(line)
                        else:
                            lines_decode_step.append(line)
                    if node_info != None:
                        if 'Times' in line or node_info.startswith('P'):
                            lines[-1] = lines[-1] + '|NODE=' + node_info + '.'
                        elif 'worker_step start' in line:
                            # profile: worker_step start:4790|1751437716.6964805|model_step=1097
                            id0 = line.find('worker_step start:')
                            id1 = line.find('|', id0)
                            id2 = line.find('|', id1 + 1)
                            d_step_key = node_info + '|' + line[id0 + len('worker_step start:'):id2]
                            worker_step = int(line[id2 + len('|model_step='):])
                            lines_decode_worker_step[d_step_key] = worker_step
                        else:
                            lines_decode_step[-1] = lines_decode_step[-1] + '|NODE=' + node_info + '.'
                            idx0 = item.find('_NODE_') + len('_NODE_')
                            idx1 = item.find('_', idx0) + 1
                            idx0 = item.find('_', idx1)
                            pid = item[idx1:idx0]
                            lines_decode_step[-1] = lines_decode_step[-1] + '|PID=' + pid + '.'

    print(f'read all files done')

    req_datas = {}  # req_id : timstamp actions
    decode_token_nums = {}  # reqid : decoder tokens
    for line in metric:
        if 'CompletionMetric' in line:
            idx0 = line.find('profile REQ_ID[') + len('profile REQ_ID[')
            idx1 = line.find(']', idx0)
            req_id = line[idx0:idx1]
            idx0 = line.find('num_completion_tokens=') + len('num_completion_tokens=')
            idx1 = line.find('.', idx0)
            output_tokens = line[idx0:idx1]
            decode_token_nums[req_id] = int(output_tokens)
    print(f'parse CompletionMetric done')

    for line in lines_decode_step:
        if 'engine_step start' in line:
            # print(f"{line=}")
            id0 = line.find('engine_step start:')
            id1 = line.find('|finish:', id0)
            start_timestamp = line[id0 + len('engine_step start:'):id1]
            start_timestamp = float(start_timestamp)
            id0 = line.find('|finish:')
            id1 = line.find('|execute time:', id0)
            finish_timestamp = line[id0 + len('|finish:'):id1]
            id0 = line.find('|execute time:')
            id1 = line.find('|seqs:', id0)
            execute_time = line[id0 + len('|execute time:'):id1]
            id0 = line.find('|seqs:')
            id1 = line.find('|tokens:', id0)
            seqs = line[id0 + len('|seqs:'):id1]
            id0 = line.find('|tokens:')
            id1 = line.find('|waiting_reqs_num_after_step=', id0)
            tokens = line[id0 + len('|tokens:'):id1]
            if int(tokens) == 0:
                continue
            id0 = line.find('|waiting_reqs_num_after_step=')
            id1 = line.find('|reqs_ids=', id0)
            waiting_num = line[id0 + len('|waiting_reqs_num_after_step='):id1]
            id0 = line.find('|reqs_ids=')
            id1 = line.find('|bs_tokens=', id0)
            reqs = line[id0 + len('|reqs_ids='):id1]
            pattern = r'[0-9]+'
            uuids_reqs = re.findall(pattern, reqs)
            # print(f"{uuids_reqs=}")
            id0 = line.find('|bs_tokens=')
            id1 = line.find('|execute_model_start_time=', id0)
            tokens_per_req = ast.literal_eval(line[id0 + len('|bs_tokens='):id1])
            id0 = line.find('|execute_model_start_time=')
            id1 = line.find('|execute_model_end_time=', id0)
            model_start_timestamp = line[id0 + len('|execute_model_start_time='):id1]
            id0 = line.find('|execute_model_end_time=')
            id1 = line.find('|execute_model_cost_time=', id0)
            model_finish_timestamp = line[id0 + len('|execute_model_end_time='):id1]
            id0 = line.find('|execute_model_cost_time=')
            id1 = line.find('|kv_cache_usage=', id0)
            model_execute_time = line[id0 + len('|execute_model_cost_time='):id1]
            id0 = line.find('|kv_cache_usage=')
            id1 = line.find('|kv_blocks_num=', id0)
            kv_cache_usage = line[id0 + len('|kv_cache_usage='):id1]
            kv_cache_usage = float(kv_cache_usage)
            id0 = line.find('|kv_blocks_num=')
            id1 = line.find('|start_free_block_num=', id0)
            kv_blocks_num = line[id0 + len('|kv_blocks_num='):id1]
            id0 = line.find('|start_free_block_num=')
            id1 = line.find('|end_free_block_num=', id0)
            start_free_block_num = line[id0 + len('|start_free_block_num='):id1]
            id0 = line.find('|end_free_block_num=')
            id1 = line.find('|cost_blocks_num=', id0)
            end_free_block_num = line[id0 + len('|end_free_block_num='):id1]
            id0 = line.find('|cost_blocks_num=')
            id1 = line.find('|engine_core_str=', id0)
            cost_blocks_num = line[id0 + len('|cost_blocks_num='):id1]
            id0 = line.find('|engine_core_str=')
            id1 = line.find('|NODE=', id0)
            engine_core_str = line[id0 + len('|engine_core_str='):id1]
            id0 = line.find('|NODE=')
            id1 = line.find('.', id0)
            node = line[id0 + len('|NODE='):id1]
            id0 = line.find('|PID=', id1)
            id1 = line.find('.', id0)
            pid = line[id0 + len('|PID='):id1]
            worker_step = decode_worker_step(node, engine_core_str, lines_decode_worker_step)
            if worker_step is None:
                print(f'not find worker_step for [{line}]')
            dict_key = node + '_' + pid
            if dict_key not in step_data_decode.keys():
                step_data_decode[dict_key] = []
            step_data_decode[dict_key].append(
                [start_timestamp, finish_timestamp, execute_time, seqs, int(tokens), waiting_num, \
                 float(model_start_timestamp), float(model_finish_timestamp), float(model_execute_time), kv_cache_usage,
                 kv_blocks_num, start_free_block_num, \
                 end_free_block_num, cost_blocks_num, worker_step])
    print(f'parse decode_step done')

    for line in lines:
        if 'engine_step start' in line:
            # print(f"{line=}")
            id0 = line.find('engine_step start:')
            id1 = line.find('|finish:', id0)
            start_timestamp = line[id0 + len('engine_step start:'):id1]
            start_timestamp = float(start_timestamp)
            id0 = line.find('|finish:')
            id1 = line.find('|execute time:', id0)
            finish_timestamp = line[id0 + len('|finish:'):id1]
            id0 = line.find('|execute time:')
            id1 = line.find('|seqs:', id0)
            execute_time = line[id0 + len('|execute time:'):id1]
            id0 = line.find('|seqs:')
            id1 = line.find('|tokens:', id0)
            seqs = line[id0 + len('|seqs:'):id1]
            id0 = line.find('|tokens:')
            id1 = line.find('|waiting_reqs_num_after_step=', id0)
            tokens = line[id0 + len('|tokens:'):id1]
            if int(tokens) == 0:
                continue
            id0 = line.find('|waiting_reqs_num_after_step=')
            id1 = line.find('|reqs_ids=', id0)
            waiting_num = line[id0 + len('|waiting_reqs_num_after_step='):id1]
            id0 = line.find('|reqs_ids=')
            id1 = line.find('|bs_tokens=', id0)
            reqs = line[id0 + len('|reqs_ids='):id1]
            pattern = r'[0-9]+'
            uuids_reqs = re.findall(pattern, reqs)
            # print(f"{uuids_reqs=}")
            id0 = line.find('|bs_tokens=')
            id1 = line.find('|execute_model_start_time=', id0)
            tokens_per_req = ast.literal_eval(line[id0 + len('|bs_tokens='):id1])
            id0 = line.find('|execute_model_start_time=')
            id1 = line.find('|execute_model_end_time=', id0)
            model_start_timestamp = line[id0 + len('|execute_model_start_time='):id1]
            id0 = line.find('|execute_model_end_time=')
            id1 = line.find('|execute_model_cost_time=', id0)
            model_finish_timestamp = line[id0 + len('|execute_model_end_time='):id1]
            id0 = line.find('|execute_model_cost_time=')
            id1 = line.find('|kv_cache_usage=', id0)
            model_execute_time = line[id0 + len('|execute_model_cost_time='):id1]
            id0 = line.find('|kv_cache_usage=')
            id1 = line.find('|kv_blocks_num=', id0)
            kv_cache_usage = line[id0 + len('|kv_cache_usage='):id1]
            kv_cache_usage = float(kv_cache_usage)
            id0 = line.find('|kv_blocks_num=')
            id1 = line.find('|start_free_block_num=', id0)
            kv_blocks_num = line[id0 + len('|kv_blocks_num='):id1]
            id0 = line.find('|start_free_block_num=')
            id1 = line.find('|end_free_block_num=', id0)
            start_free_block_num = line[id0 + len('|start_free_block_num='):id1]
            id0 = line.find('|end_free_block_num=')
            id1 = line.find('|cost_blocks_num=', id0)
            end_free_block_num = line[id0 + len('|end_free_block_num='):id1]
            id0 = line.find('|cost_blocks_num=')
            id1 = line.find('|engine_core_str=', id0)
            cost_blocks_num = line[id0 + len('|cost_blocks_num='):id1]
            id0 = line.find('|engine_core_str=')
            id1 = line.find('|NODE=', id0)
            engine_core_str = line[id0 + len('|engine_core_str='):id1]
            id0 = line.find('|NODE=')
            id1 = line.find('.', id0)
            node = line[id0 + len('|NODE='):id1]
            step_data.append([start_timestamp, finish_timestamp, execute_time, seqs, tokens, waiting_num, uuids_reqs, \
                              tokens_per_req, model_start_timestamp, model_finish_timestamp, model_execute_time, \
                              kv_cache_usage, kv_blocks_num, start_free_block_num, end_free_block_num, cost_blocks_num,
                              node])
        if 'Times' in line:
            # print(f"{line=}")
            first_index = line.index('profile REQ_ID[')
            second_index = line.index(']', first_index)
            fifth_index = line.index(' action:', second_index)
            sixth_index = line.index('.', fifth_index)
            seventh_index = line.index('Timestamp ', sixth_index)
            last_index = line.index('|', seventh_index)

            # time_str = line[third_index + len('time['):forth_index].strip()
            action = line[fifth_index + len(' action:'):sixth_index].strip()
            timestamp = float(line[seventh_index + len('Timestamp '):last_index].strip())
            first_index = first_index + len('profile REQ_ID[')
            req_id = line[first_index:second_index]
            if req_id == "0":
                continue
            d_node = None
            p_node = None
            if 'NODE=D' in line:
                id0 = line.find('NODE=')
                id1 = line.find('.', id0)
                d_node = line[id0 + len('NODE='):id1]
            elif 'NODE=P' in line:
                id0 = line.find('NODE=')
                id1 = line.find('.', id0)
                p_node = line[id0 + len('NODE='):id1]

            if 'Add need pullling sequence' in action:
                values = action.split('|')
                action = values[0]
                waiting_pull_len_info = values[1].split('=')
                # print(f"{action=}, {timestamp=}, {waiting_pull_len_info[0]=}, {waiting_pull_len_info[1]=}")
                if req_id not in req_datas:
                    req_datas[req_id] = {action: [timestamp], waiting_pull_len_info[0]: [waiting_pull_len_info[1]]}
                else:
                    req_datas[req_id].update({action: [timestamp]})
                    req_datas[req_id].update({waiting_pull_len_info[0]: [waiting_pull_len_info[1]]})
            elif 'p_4 tokennizer to sche' in action:
                values = action.split('|')
                action = values[0]
                seq_len_info = values[1].split('=')
                if req_id not in req_datas:
                    req_datas[req_id] = {action: [timestamp], seq_len_info[0]: [seq_len_info[1]]}
                else:
                    req_datas[req_id].update({action: [timestamp]})
                    req_datas[req_id].update({seq_len_info[0]: [int(seq_len_info[1])]})

            else:
                if action not in action_dict.keys():
                    continue
                if req_id not in req_datas:
                    req_datas[req_id] = {action: []}
                    req_datas[req_id][action].append(timestamp)
                else:
                    if action not in req_datas[req_id]:
                        req_datas[req_id].update({action: []})
                    req_datas[req_id][action].append(timestamp)
                    if d_node is not None:
                        req_datas[req_id].update({'DNode': [d_node]})
                    if p_node is not None:
                        req_datas[req_id].update({'PNode': [p_node]})
    print(f'parse and prefill action timestamp done')

    result = {}
    for req_id, data in req_datas.items():
        result[req_id] = {}
        for miss_key in action_dict.keys() - data.keys():
            result[req_id].update({miss_key: 'NA'})
        for action, time_list in data.items():
            result[req_id][action] = min(time_list)
    for req_id, data in result.items():
        if req_id in decode_token_nums.keys():
            result[req_id].update({"decode token number": decode_token_nums[req_id]})
        else:
            result[req_id].update({"decode token number": 'NA'})

    return step_data, step_data_decode, result


def save_to_time_analysis_detail(result, output_path="time_analysis.xlsx"):
    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "time_analysis"
    data_ws.append(_sgl_title())
    for req_id, data in result.items():
        data_line = [req_id]

        # data_line.append("NA") # Prefill id
        data_line.append(data['Seq len'])
        data_line.append(data['decode token number'])
        data_line.append(data['PNode'])
        data_line.append(data['DNode'])

        data_line.append(data['Start to schedule'])
        data_line.append(data['Finish to choose device and add request'])
        data_line.append(data['Start to send request to pd api server'])

        data_line.append(data['p_1 prefill api server收到请求'])
        data_line.append(data['p_2 触发engine处理请求'])
        data_line.append(data['p_3 engine开始tokennizer'])
        data_line.append(data['p_3 engine结束tokennizer'])
        data_line.append(data['p_4 tokennizer to sche'])
        data_line.append(data['p_5 P侧添加到bootstrap队列之后'])
        data_line.append(data['P侧握手完成'])
        data_line.append(data['p_6 握手完成加到waiting队列'])
        data_line.append(data['p_7 组batch完成 开始run batch'])
        data_line.append(data['p_8 Push a new batch to the input queue'])
        data_line.append(data['p_9 开始执行Run forward'])
        data_line.append(data['p_9 结束执行Run forward'])
        data_line.append(data['p_13 开始发送kv cache'])
        data_line.append(data['p_14 完成发送kv cache'])
        data_line.append(data['p_15 P侧释放KV'])
        data_line.append(data['p_16 client收到输出并入队'])
        data_line.append(data['p_17 client出队'])
        data_line.append(data['p_18 api server收到请求准备返回'])

        data_line.append(data['Get response from pd api server'])
        data_line.append(data['Finish to chosse device and start decode generate'])
        data_line.append(data['Finish to decode'])

        data_line.append(data['d_1 decode api server收到请求'])
        data_line.append(data['d_2 触发engine处理请求'])
        data_line.append(data['d_3 engine开始tokennizer'])
        data_line.append(data['d_3 engine结束tokennizer'])
        data_line.append(data['d_4 tokennizer to sche'])
        data_line.append(data['d_5 scheduler开始处理请求'])
        data_line.append(data['d_6 D侧添加到prealloc_queue队列之后'])
        data_line.append(data['d_8 开始分配kv缓存'])
        data_line.append(data['d_9 d侧开始握手'])
        data_line.append(data['d_7 Add need pullling sequence'])
        data_line.append(data['waiting_pull_len'])
        data_line.append(data['d_11 d侧收到kv传输完成通知'])
        data_line.append(data['d_12 轮询到状态 kv cache传输完成 开始加入到waiting队列'])
        data_line.append(data['d_18 触发首个decode token执行'])
        data_line.append(data['d_19 decoder返回第一个token'])
        data_line.append(data['d_20 decoder返回第二个token'])
        data_line.append(data['d_21 api server收到推理结果'])

        data_ws.append(data_line)
    wb.save(output_path)
    print(f"已保存结果到 {output_path}")


def save_to_engine_step_detail_prefill(result, data_ws, start_time):
    engine_step_title = ['Engine step开始时间', 'Engine step结束时间', '执行时间(ms)', 'Seq数量', 'Token数量',
                         '处理完成后waiting队列长度', 'reqids', 'tokens per req', '模型开始时间', '模型结束时间',
                         '模型执行时间(ms)', \
                         'kv usage', 'kv block总数', 'step初始空闲block数', 'step结束空闲block数',
                         'step新增使用block数', '节点']
    data_ws.title = "engine_step"

    # print(f"排序前：{result}")
    # 按照时间戳排序
    sorted_result = sorted(result, key=lambda x: x[0])
    idx = 0
    for r in sorted_result:
        if r[0] < start_time:
            idx += 1
    # print(f"排序后：{result}")
    # 删除小于时间戳的
    del sorted_result[:idx]

    data_ws.append(engine_step_title)  #
    for r in sorted_result:
        # print(f"node is {r[-1]}")
        processed_row = []
        for cell in r:
            processed_row.append(str(cell))
        data_ws.append(processed_row)


def save_to_engine_step_detail_decode(result_decode, decode_data_ws):
    engine_step_title = ['节点_die', 'Engine step开始时间', 'Engine step结束时间', '执行时间(ms)', 'Seq数量',
                         'Token数量', '处理完成后waiting队列长度', \
                         '模型开始时间', '模型结束时间', '模型执行时间(ms)', 'kv usage', 'kv block总数',
                         'step初始空闲block数', 'step结束空闲block数', 'step新增使用block数', \
                         'step轮次']
    decode_data_ws.append(engine_step_title)  #
    for key, value in result_decode.items():
        for r in value:
            # print(f"node is {r[-1]}")
            processed_row = []
            processed_row.append(key)
            for cell in r:
                if cell is None:
                    processed_row.append(None)
                else:
                    processed_row.append(str(cell))
            decode_data_ws.append(processed_row)


def save_to_engine_step_decode_die_load(result_decode, decode_die_load_ws):
    '''
    # 第一行写入字段名
    decode_die_load_ws.append(list(result_decode.keys()))
    # 后续行写入数据（转置列数据为行）
    for row in zip(*result_decode.values()):
        decode_die_load_ws.append(row)
    '''
    out_dict = {}
    # print(f"{result_decode}")
    for key, value in result_decode.items():
        step_key = key + '_step轮次'
        step = []
        time_key = key + '_模型开始时间'
        time = []
        seqs_key = key + '_Seq数量'
        seqs = []
        tokens_key = key + '_Token数量'
        tokens = []
        waiting_key = key + '_处理完成后waiting队列长度'
        waiting = []
        model_exec_key = key + '_模型执行时间(ms)'
        model_exec = []
        # kv_num_key = key + '_kv block总数'
        # kv_num = []
        kv_free_key = key + '_step结束空闲block数'
        kv_free = []
        kv_cost_key = key + '_step新增使用block数'
        kv_cost = []
        for r in value:
            step.append(r[14])
            time.append(r[6])
            seqs.append(r[3])
            tokens.append(r[4])
            waiting.append(r[5])
            model_exec.append(r[8])
            # kv_num.append(r[10])
            kv_free.append(r[12])
            kv_cost.append(r[13])
        out_dict[step_key] = step
        out_dict[time_key] = time
        out_dict[seqs_key] = seqs
        out_dict[tokens_key] = tokens
        out_dict[waiting_key] = waiting
        out_dict[model_exec_key] = model_exec
        # out_dict[kv_num_key] = kv_num
        out_dict[kv_free_key] = kv_free
        out_dict[kv_cost_key] = kv_cost
    # 第一行写入字段名
    decode_die_load_ws.append(list(out_dict.keys()))
    # 后续行写入数据（转置列数据为行）
    for row in zip(*out_dict.values()):
        decode_die_load_ws.append(row)


def save_to_engine_step_decode_die_time(result_decode, steps, ws):
    # 计算公共起始时间
    s_time = 4906457181.437  # 2125-06-25 00:44:51.437
    for value in result_decode.values():
        # print(f"{value[0][0]=}")
        if value[0][6] is not None:
            s_time = min(s_time, value[0][6])
    print(f"{s_time=}")

    out_dict = {}
    tokens_plot = {}
    tokens_max_plot = {}
    step_token_dict = {}
    tokens_max = []
    tokens_mean = []
    max_div_mean = []
    for key, value in result_decode.items():
        for r in value:
            if r[4] is None:
                continue
            if r[14] not in step_token_dict.keys():
                step_token_dict[r[14]] = []
            step_token_dict[r[14]].append(r[4])
    # print(f"{step_token_dict=}")
    for _ in steps:
        max = np.max(step_token_dict[_])
        mean = np.mean(step_token_dict[_])
        tokens_max.append(max)
        tokens_mean.append(mean)
        max_div_mean.append(float(max / mean))
    out_dict['step轮次'] = steps
    out_dict['Token数量(Max)'] = tokens_max
    out_dict['Token数量(Avg)'] = tokens_mean
    out_dict['Token数量(Max/Avg)'] = max_div_mean
    # print(f"Token数量统计: {out_dict=}")

    for key, value in result_decode.items():
        time_key = key + '_模型开始时间'
        time = []
        tokens_key = key + '_Token数量'
        tokens = []
        for r in value:
            if r[6] is not None:
                time.append(r[6] - s_time)
            else:
                time.append(None)
            tokens.append(r[4])
        out_dict[time_key] = time
        out_dict[tokens_key] = tokens
        tokens_plot[key] = {'x': time, 'y': tokens}
    # print(f"{out_dict=}")
    # print(f"{tokens_plot=}")
    # print(f"{tokens_max_plot=}")
    # 第一行写入字段名
    ws.append(list(out_dict.keys()))
    # 后续行写入数据（转置列数据为行）
    for row in zip(*out_dict.values()):
        ws.append(row)

    data = list(ws.values)
    if not data:
        raise ValueError("工作表中没有数据。")
    df = pd.DataFrame(data[1:], columns=data[0])

    # 输出示例
    # print("数据前几行：")
    # print(df.head())
    # print("\n数据维度：", df.shape)

    # 画“Token数量”散点图
    chart_tokens = ScatterChart()
    chart_tokens.title = 'step轮次 vs Token数量'
    chart_tokens.style = 13
    chart_tokens.x_axis.title = 'step轮次'
    chart_tokens.y_axis.title = 'Token数量'

    # colors = ["FF0000", "00FF00", "0000FF"]  # 红绿蓝
    colors = ['#ce2e2e', '#c73324', '#b05e4b', '#cb714d', '#da631b', '#f49d51', '#d49f5a', '#e7bf72', '#d9ba5b',
              '#c9b33e', \
              '#e2d83f', '#f1f746', '#a8b936', '#b6e021', '#91c421', '#81b139', '#77a647', '#6ca744', '#69ea2e',
              '#72f44e', \
              '#26a714', '#2ae224', '#1af828', '#64fb7c', '#30b451', '#2bd766', '#1acc68', '#2caf72', '#3cd49b',
              '#3db594', \
              '#1dc2a3', '#5cddd1', '#4caaaa', '#219ba8', '#6dcee5', '#4c96b2', '#237aae', '#41a5fd', '#2f5d99',
              '#2f5aac', \
              '#3a63de', '#1a36ce', '#2430e8', '#5c59c4', '#5b45f4', '#3b2099', '#815fcc', '#662bbc', '#6619b3',
              '#a051d6', \
              '#9753b6', '#be5cd9', '#d71df1', '#ac35b0', '#e158d9', '#af3f9e', '#c459a9', '#f74abb', '#a82870',
              '#db3c86', \
              '#e74582', '#ec7193', '#f6385c', '#ca1526']
    # colors = generate_distinct_colors(n=64)
    for idx, (model_name, model_data) in enumerate(tokens_plot.items()):
        x_col = 'step轮次'
        y_col = model_name + '_Token数量'
        # print(f"{x_col=}, {y_col=}")
        # print(f"{df.columns.get_loc(x_col)=}, {df.columns.get_loc(y_col)=}")
        x_range = Reference(ws, min_col=df.columns.get_loc(x_col) + 1, min_row=2, max_row=len(steps) + 1)
        y_range = Reference(ws, min_col=df.columns.get_loc(y_col) + 1, min_row=2, max_row=len(model_data['y']) + 1)

        s = Series(y_range, x_range, title=model_name)
        marker = Marker()
        marker.symbol = 'circle'
        marker.size = 7
        # 设置空心效果和边框颜色
        # marker.spPr = marker.spPr or Marker().spPr
        # marker.spPr.solidFill = None

        # # 为不同类别设置不同颜色
        # color = colors[idx]
        # marker.spPr.ln = marker.spPr.ln or Marker().spPr.ln
        # marker.spPr.ln.solidFill = ColorChoice(srgbClr=color)
        # marker.spPr.ln.w = 12700

        # 移除连接线（仅保留散点）
        s.graphicalProperties.line.noFill = True
        s.marker = marker
        chart_tokens.series.append(s)
    ws.add_chart(chart_tokens, "L6")

    # 画“Token数量(Max/Avg)”散点图
    chart_tokens_max = ScatterChart()
    chart_tokens_max.title = "step轮次 vs Token数量(Max/Avg)"
    chart_tokens_max.style = 13
    chart_tokens_max.x_axis.title = "step轮次"
    chart_tokens_max.y_axis.title = "Token数量(Max/Avg)"
    chart_tokens_max.legend = None

    x_col = 'step轮次'
    y_col = 'Token数量(Max/Avg)'
    # print(f"{x_col=}, {y_col=}")
    # print(f"{df.columns.get_loc(x_col)=}, {df.columns.get_loc(y_col)=}")
    x_range = Reference(ws, min_col=df.columns.get_loc(x_col) + 1, min_row=2, max_row=len(steps) + 1)
    y_range = Reference(ws, min_col=df.columns.get_loc(y_col) + 1, min_row=2, max_row=len(steps) + 1)

    s = Series(y_range, x_range, title='Token数量(Max/Avg)')
    marker = Marker()
    marker.symbol = 'circle'
    marker.size = 7
    # 设置空心效果和边框颜色
    # marker.spPr = marker.spPr or Marker().spPr
    # marker.spPr.solidFill = None

    # # 为不同类别设置不同颜色
    # color = colors[idx]
    # marker.spPr.ln = marker.spPr.ln or Marker().spPr.ln
    # marker.spPr.ln.solidFill = ColorChoice(srgbClr=color)
    # marker.spPr.ln.w = 12700

    # 移除连接线（仅保留散点）
    s.graphicalProperties.line.noFill = True
    s.marker = marker
    chart_tokens_max.series.append(s)
    ws.add_chart(chart_tokens_max, "L25")


def save_to_engine_step_detail(result, result_decode, start_time, output_path="engine_step.xlsx"):
    wb = Workbook()
    data_ws = wb.active

    save_to_engine_step_detail_prefill(result, data_ws, start_time)
    print(f"写入 {data_ws.title} 完成")

    # print(f"{result_decode=}")

    # result_decode中每一项按照时间进行排序
    for key, value in result_decode.items():
        result_decode[key] = sorted(value, key=lambda x: x[0])
    # print(f"sorted:{result_decode=}")
    # 在时间排序基础上，将时间戳前面的删除
    # del data[:2]  # 删除前两个元素
    for key, value in result_decode.items():
        idx = 0
        for r in value:
            if r[0] < start_time:
                idx += 1
        del value[:idx]

    '''
    # result_decode中每一项按照step轮次进行排序
    for key, value in result_decode.items():
        result_decode[key] = sorted(value, key=lambda x: x[14]) # 当前编号是14
    # print(f"sorted:{result_decode=}")
    # 将前面的xxx轮次删除
    if start_step != 0:
        for key, value in result_decode.items():
            del value[:start_step]
    '''

    decode_data_ws = wb.create_sheet(title="engine_step_decode")
    save_to_engine_step_detail_decode(result_decode, decode_data_ws)
    print(f"写入 {decode_data_ws.title} 完成")

    # 获取所有step编号，为每个die补齐不存在的step
    steps = []
    die_steps_dict = {}
    die_steps = []
    for key, value in result_decode.items():
        die_steps = []
        for i in range(len(value)):
            if len(die_steps) != 0 and value[i][14] == die_steps[-1]:
                print(f'step {value[i][14]} duplicuted in {key}')
            else:
                die_steps.append(value[i][14])
            if not value[i][14] in steps:
                steps.append(value[i][14])
        die_steps_dict[key] = die_steps
    steps.sort()
    # print(f"{steps=}")

    find = False
    for step in steps:
        for key, value in die_steps_dict.items():
            if step not in value:
                find = False
                none_list = [None] * 15
                none_list[14] = step
                for i in range(len(result_decode[key])):
                    if result_decode[key][i][14] > step:
                        find = True
                        result_decode[key].insert(i, none_list)
                        break
                if not find:
                    result_decode[key].append(none_list)
    # print(f"{die_steps_dict=}")
    # print(f"{result_decode=}")

    decode_die_load_ws = wb.create_sheet(title="Decode_die_load")
    save_to_engine_step_decode_die_load(result_decode, decode_die_load_ws)
    print(f"写入 {decode_die_load_ws.title} 完成")

    decode_die_load_time_ws = wb.create_sheet(title="Decode_die_time")
    save_to_engine_step_decode_die_time(result_decode, steps, decode_die_load_time_ws)
    print(f"写入 {decode_die_load_time_ws.title} 完成")

    wb.save(output_path)
    print(f"已保存结果到 {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("请提供解析目录和时间戳")  # 2025-06-25 00:44:51.437
        sys.exit(-1)

    print(f"log path: {sys.argv[1]}, {sys.argv[2]}")
    folder_path = sys.argv[1]
    timestamp_str = sys.argv[2]
    dt_obj = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S.%f")
    start_time = dt_obj.timestamp()
    step_data, step_data_decode, result = parse_file(folder_path)
    print(f"{start_time=}")
    print(f"{len(step_data)=}")
    print(f"{len(step_data_decode)=}")
    print(f"{len(result)=}")
    save_to_time_analysis_detail(result)
    save_to_engine_step_detail(step_data, step_data_decode, start_time)